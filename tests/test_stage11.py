import argparse
import csv
import json
import os
import unittest
from datetime import datetime, timezone
from pathlib import Path
from unittest.mock import MagicMock, patch

import openpyxl

from app.database import DatabaseError
from app.exporter import ExportService
from app.models import LinkedInJob
from app.repository import JobRepository
from main import main


class Stage11Tests(unittest.TestCase):
    def setUp(self) -> None:
        self.test_dir = Path("data/test_exports")
        self.test_dir.mkdir(parents=True, exist_ok=True)
        self.csv_path = self.test_dir / "jobs.csv"
        self.excel_path = self.test_dir / "jobs.xlsx"
        self.json_path = self.test_dir / "jobs.json"

    def tearDown(self) -> None:
        for path in (self.csv_path, self.excel_path, self.json_path):
            if path.exists():
                try:
                    path.unlink()
                except Exception:
                    pass
        if self.test_dir.exists():
            try:
                self.test_dir.rmdir()
            except Exception:
                pass

    def _make_dummy_jobs(self) -> list[LinkedInJob]:
        return [
            LinkedInJob(
                job_title="Python Engineer",
                company_name="Acme Corp",
                company_url="https://acme.example.com",
                linkedin_job_url="https://www.linkedin.com/jobs/view/100",
                location="Remote",
                workplace_type="REMOTE",
                easy_apply=True,
                posted_date=datetime(2026, 1, 1, tzinfo=timezone.utc),
                scraped_timestamp=datetime(2026, 7, 15, tzinfo=timezone.utc),
                skills=["Python", "SQL"],
                raw_json={"jobId": "100"}
            ),
            LinkedInJob(
                job_title="Backend Developer",
                company_name="Tech Solutions",
                linkedin_job_url="https://www.linkedin.com/jobs/view/200",
                location="New York",
                workplace_type="ONSITE",
                easy_apply=False,
                posted_date=datetime(2026, 2, 1, tzinfo=timezone.utc),
                scraped_timestamp=datetime(2026, 7, 16, tzinfo=timezone.utc),
                skills=["Go", "Postgres"],
                raw_json={"jobId": "200"}
            )
        ]

    @patch("app.repository.JobRepository.get_all_jobs")
    def test_csv_export(self, mock_get_all) -> None:
        mock_get_all.return_value = self._make_dummy_jobs()
        service = ExportService()
        
        service.export_csv(self.csv_path)
        self.assertTrue(self.csv_path.exists())
        
        with open(self.csv_path, mode="r", newline="", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            rows = list(reader)
            
        self.assertEqual(len(rows), 2)
        self.assertEqual(rows[0]["job_title"], "Python Engineer")
        self.assertEqual(rows[0]["company_name"], "Acme Corp")
        self.assertEqual(rows[0]["easy_apply"], "True")
        self.assertEqual(rows[1]["job_title"], "Backend Developer")
        self.assertEqual(rows[1]["company_name"], "Tech Solutions")
        self.assertEqual(rows[1]["easy_apply"], "False")

    @patch("app.repository.JobRepository.get_all_jobs")
    def test_excel_export(self, mock_get_all) -> None:
        mock_get_all.return_value = self._make_dummy_jobs()
        service = ExportService()
        
        service.export_excel(self.excel_path)
        self.assertTrue(self.excel_path.exists())
        
        wb = openpyxl.load_workbook(self.excel_path)
        ws = wb.active
        self.assertEqual(ws.title, "Jobs")
        self.assertEqual(ws.freeze_panes, "A2")
        
        # Header bold verification
        self.assertTrue(ws.cell(row=1, column=1).font.bold)
        
        # Row verification
        self.assertEqual(ws.cell(row=2, column=1).value, "Python Engineer")
        self.assertEqual(ws.cell(row=2, column=2).value, "Acme Corp")
        self.assertEqual(ws.cell(row=3, column=1).value, "Backend Developer")
        
        # Column dimension verify
        col_letter = openpyxl.utils.get_column_letter(1)
        self.assertGreater(ws.column_dimensions[col_letter].width, 0)

    @patch("app.repository.JobRepository.get_all_jobs")
    def test_json_export(self, mock_get_all) -> None:
        mock_get_all.return_value = self._make_dummy_jobs()
        service = ExportService()
        
        service.export_json(self.json_path)
        self.assertTrue(self.json_path.exists())
        
        with open(self.json_path, mode="r", encoding="utf-8") as f:
            data = json.load(f)
            
        self.assertEqual(len(data), 2)
        self.assertEqual(data[0]["job_title"], "Python Engineer")
        self.assertEqual(data[0]["skills"], ["Python", "SQL"])
        self.assertEqual(data[0]["raw_json"], {"jobId": "100"})

    @patch("app.repository.JobRepository.get_statistics")
    def test_statistics(self, mock_get_stats) -> None:
        dummy_stats = {
            "total_jobs": 100,
            "total_companies": 10,
            "total_countries": 2,
            "remote_jobs": 60,
            "hybrid_jobs": 30,
            "onsite_jobs": 10,
            "easy_apply_jobs": 40,
            "latest_scrape_date": datetime(2026, 7, 15, tzinfo=timezone.utc),
            "oldest_scrape_date": datetime(2026, 1, 1, tzinfo=timezone.utc),
            "duplicate_count": 5
        }
        mock_get_stats.return_value = dummy_stats
        service = ExportService()
        
        stats = service.get_statistics()
        self.assertEqual(stats["total_jobs"], 100)
        self.assertEqual(stats["remote_jobs"], 60)
        self.assertEqual(stats["duplicate_count"], 5)

    @patch("app.repository.get_connection")
    def test_empty_database_statistics(self, mock_get_connection) -> None:
        # Mock cursor returning zeros/nones for empty DB queries
        mock_conn = MagicMock()
        mock_cursor = MagicMock()
        mock_get_connection.return_value.__enter__.return_value = mock_conn
        mock_conn.cursor.return_value.__enter__.return_value = mock_cursor
        
        # Configure select queries outputs sequentially
        mock_cursor.fetchone.side_effect = [
            (0,), # total_jobs
        ]
        
        repository = JobRepository()
        stats = repository.get_statistics()
        
        self.assertEqual(stats["total_jobs"], 0)
        self.assertEqual(stats["total_companies"], 0)
        self.assertIsNone(stats["latest_scrape_date"])

    def test_invalid_output_path(self) -> None:
        service = ExportService()
        invalid_path = Path("invalid_dir_xyz/jobs.csv")
        
        with self.assertRaises(FileNotFoundError):
            service.export_csv(invalid_path)
            
        with self.assertRaises(FileNotFoundError):
            service.export_excel(invalid_path)
            
        with self.assertRaises(FileNotFoundError):
            service.export_json(invalid_path)

    @patch("app.repository.get_connection")
    def test_repository_get_all_jobs(self, mock_get_connection) -> None:
        mock_conn = MagicMock()
        mock_cursor = MagicMock()
        mock_get_connection.return_value.__enter__.return_value = mock_conn
        mock_conn.cursor.return_value.__enter__.return_value = mock_cursor
        
        dummy_row = [
            "Staff Engineer", "Innovate Inc", "https://innovate.example.com",
            "https://www.linkedin.com/jobs/view/300", "300", "San Francisco", "US",
            "HYBRID", "FULL_TIME", "SENIOR", "$150k", "USD", "Description snippet",
            "Job summary text", ["Python", "AWS"], "Software", "Health insurance",
            "Recruiter Name", "https://recruiter.example.com", "https://logo.example.com",
            "100-500", "https://apply.example.com", True,
            datetime(2026, 3, 1, tzinfo=timezone.utc), datetime(2026, 7, 17, tzinfo=timezone.utc),
            {"jobId": "300"}
        ]
        mock_cursor.fetchall.return_value = [dummy_row]
        
        repository = JobRepository()
        jobs = repository.get_all_jobs(limit=10, offset=5)
        
        self.assertEqual(len(jobs), 1)
        self.assertEqual(jobs[0].job_title, "Staff Engineer")
        self.assertEqual(jobs[0].workplace_type, "HYBRID")
        self.assertTrue(jobs[0].easy_apply)
        self.assertEqual(jobs[0].raw_json, {"jobId": "300"})

    @patch("main.parse_arguments")
    @patch("app.config_validator.load_environment")
    @patch("app.config_validator.validate_required_env")
    @patch("app.config_validator.build_config")
    @patch("app.config_validator.verify_postgres_connection")
    @patch("app.exporter.ExportService.get_statistics")
    def test_cli_stats_command_success(
        self, mock_stats, mock_verify, mock_build, mock_validate, mock_load, mock_parse
    ) -> None:
        mock_parse.return_value = argparse.Namespace(
            stats=True,
            export=None,
            output=None,
            schedule=False,
            interval=None,
            run_once=False
        )
        mock_stats.return_value = {
            "total_jobs": 5, "total_companies": 2, "total_countries": 1,
            "remote_jobs": 3, "hybrid_jobs": 1, "onsite_jobs": 1,
            "easy_apply_jobs": 2, "latest_scrape_date": None, "oldest_scrape_date": None,
            "duplicate_count": 0
        }
        with patch("sys.stdout"), patch("sys.stderr"):
            exit_code = main()
        self.assertEqual(exit_code, 0)

    @patch("main.parse_arguments")
    @patch("app.config_validator.load_environment")
    @patch("app.config_validator.validate_required_env")
    @patch("app.config_validator.build_config")
    @patch("app.config_validator.verify_postgres_connection")
    @patch("app.exporter.ExportService.export_csv")
    def test_cli_export_csv_command_success(
        self, mock_csv, mock_verify, mock_build, mock_validate, mock_load, mock_parse
    ) -> None:
        mock_parse.return_value = argparse.Namespace(
            stats=False,
            export="csv",
            output="data/test_exports/output.csv",
            schedule=False,
            interval=None,
            run_once=False
        )
        with patch("sys.stdout"), patch("sys.stderr"):
            exit_code = main()
        self.assertEqual(exit_code, 0)
        mock_csv.assert_called_once_with("data/test_exports/output.csv")


if __name__ == "__main__":
    unittest.main()
