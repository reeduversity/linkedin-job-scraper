import csv
import json
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Font

from app.models import LinkedInJob
from app.repository import JobRepository


class ExportService:
    """Production-ready data export service for LinkedIn job listings."""

    def __init__(self, repository: JobRepository | None = None) -> None:
        self.repository = repository or JobRepository()

    def get_statistics(self) -> dict[str, Any]:
        """Fetch job stats from the database."""
        return self.repository.get_statistics()

    def _prepare_data(self) -> list[dict[str, Any]]:
        """Fetch all jobs from the repository and normalize them for export."""
        jobs = self.repository.get_all_jobs()
        rows = []
        for job in jobs:
            rows.append({
                "job_title": job.job_title or "",
                "company_name": job.company_name or "",
                "company_url": job.company_url or "",
                "linkedin_job_url": job.linkedin_job_url or "",
                "job_id": job.job_id or "",
                "location": job.location or "",
                "country": job.country or "",
                "workplace_type": job.workplace_type or "",
                "employment_type": job.employment_type or "",
                "experience_level": job.experience_level or "",
                "salary": job.salary or "",
                "currency": job.currency or "",
                "description": job.description or "",
                "job_summary": job.job_summary or "",
                "skills": ", ".join(job.skills) if job.skills else "",
                "industry": job.industry or "",
                "benefits": job.benefits or "",
                "recruiter": job.recruiter or "",
                "recruiter_url": job.recruiter_url or "",
                "company_logo": job.company_logo or "",
                "company_size": job.company_size or "",
                "application_url": job.application_url or "",
                "easy_apply": str(job.easy_apply) if job.easy_apply is not None else "",
                "posted_date": job.posted_date.isoformat() if job.posted_date else "",
                "scraped_timestamp": job.scraped_timestamp.isoformat() if job.scraped_timestamp else "",
                "source_type": job.source_type or "",
                "post_url": job.post_url or "",
                "post_text": job.post_text or "",
                "post_author_name": job.post_author_name or "",
                "post_author_profile_url": job.post_author_profile_url or "",
                "post_author_role": job.post_author_role or "",
                "poster_designation": job.poster_designation or "",
                "poster_role_category": job.poster_role_category or "",
                "hiring_confidence_score": job.hiring_confidence_score,
                "detection_method": job.detection_method or "",
                "extraction_method": job.extraction_method or "",
                "extraction_quality": job.extraction_quality or "",
                "image_url": job.image_url or "",
                "ocr_text": job.ocr_text or "",
                "ocr_confidence": job.ocr_confidence,
                "ocr_processed": job.ocr_processed,
                "ocr_extraction_status": job.ocr_extraction_status or "",
                "hashtags": ", ".join(job.hashtags) if job.hashtags else "",
                "application_method": job.application_method or "",
                "application_methods": ", ".join(job.application_methods) if job.application_methods else "",
                "application_email": job.application_email or "",
                "application_emails": ", ".join(job.application_emails) if job.application_emails else "",
                "application_platform": job.application_platform or "",
                "application_urls": ", ".join(job.application_urls) if job.application_urls else "",
                "application_form_url": job.application_form_url or "",
                "application_url_type": job.application_url_type or "",
            })
        return rows

    def export_csv(self, path: str | Path) -> None:
        """Export stored jobs to CSV format."""
        path = Path(path)
        if not path.parent.exists():
            raise FileNotFoundError(f"Output directory does not exist: {path.parent}")

        rows = self._prepare_data()
        headers = [
            "job_title", "company_name", "company_url", "linkedin_job_url", "job_id",
            "location", "country", "workplace_type", "employment_type", "experience_level",
            "salary", "currency", "description", "job_summary", "skills", "industry",
            "benefits", "recruiter", "recruiter_url", "company_logo", "company_size",
            "application_url", "easy_apply", "posted_date", "scraped_timestamp",
            "source_type", "post_url", "post_text", "post_author_name", "post_author_profile_url",
            "post_author_role", "poster_designation", "poster_role_category",
            "hiring_confidence_score", "detection_method", "extraction_method", "extraction_quality",
            "image_url", "ocr_text", "ocr_confidence", "ocr_processed", "ocr_extraction_status",
            "hashtags", "application_method", "application_methods", "application_email",
            "application_emails", "application_platform", "application_urls",
            "application_form_url", "application_url_type"
        ]

        try:
            with open(path, mode="w", newline="", encoding="utf-8") as f:
                writer = csv.DictWriter(f, fieldnames=headers)
                writer.writeheader()
                for row in rows:
                    writer.writerow({k: v for k, v in row.items() if k in headers})
        except PermissionError as exc:
            raise PermissionError(f"Permission denied to write CSV file: {path}") from exc

    def export_excel(self, path: str | Path) -> None:
        """Export stored jobs to Excel format (.xlsx)."""
        path = Path(path)
        if not path.parent.exists():
            raise FileNotFoundError(f"Output directory does not exist: {path.parent}")

        rows = self._prepare_data()
        headers = [
            "job_title", "company_name", "company_url", "linkedin_job_url", "job_id",
            "location", "country", "workplace_type", "employment_type", "experience_level",
            "salary", "currency", "description", "job_summary", "skills", "industry",
            "benefits", "recruiter", "recruiter_url", "company_logo", "company_size",
            "application_url", "easy_apply", "posted_date", "scraped_timestamp",
            "source_type", "post_url", "post_text", "post_author_name", "post_author_profile_url",
            "post_author_role", "poster_designation", "poster_role_category",
            "hiring_confidence_score", "detection_method", "extraction_method", "extraction_quality",
            "image_url", "ocr_text", "ocr_confidence", "ocr_processed", "ocr_extraction_status",
            "hashtags", "application_method", "application_methods", "application_email",
            "application_emails", "application_platform", "application_urls",
            "application_form_url", "application_url_type"
        ]

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Jobs"

        # Freeze first row
        ws.freeze_panes = "A2"

        # Bold headers
        bold_font = Font(bold=True)
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = bold_font

        # Write data rows
        for row_num, row_data in enumerate(rows, 2):
            for col_num, header in enumerate(headers, 1):
                ws.cell(row=row_num, column=col_num, value=row_data.get(header, ""))

        # Auto-size columns based on maximum content length
        for col in ws.columns:
            max_len = 0
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            for cell in col:
                val_str = str(cell.value or "")
                if len(val_str) > max_len:
                    max_len = len(val_str)
            ws.column_dimensions[col_letter].width = min(max(max_len + 3, 10), 50)

        try:
            wb.save(path)
        except PermissionError as exc:
            raise PermissionError(f"Permission denied to write Excel file: {path}") from exc

    def export_json(self, path: str | Path) -> None:
        """Export stored jobs to JSON format, preserving raw_json correctly."""
        path = Path(path)
        if not path.parent.exists():
            raise FileNotFoundError(f"Output directory does not exist: {path.parent}")

        jobs = self.repository.get_all_jobs()
        json_data = []
        for job in jobs:
            json_data.append({
                "job_title": job.job_title or "",
                "company_name": job.company_name or "",
                "company_url": job.company_url or "",
                "linkedin_job_url": job.linkedin_job_url or "",
                "job_id": job.job_id or "",
                "location": job.location or "",
                "country": job.country or "",
                "workplace_type": job.workplace_type or "",
                "employment_type": job.employment_type or "",
                "experience_level": job.experience_level or "",
                "salary": job.salary or "",
                "currency": job.currency or "",
                "description": job.description or "",
                "job_summary": job.job_summary or "",
                "skills": job.skills,
                "industry": job.industry or "",
                "benefits": job.benefits or "",
                "recruiter": job.recruiter or "",
                "recruiter_url": job.recruiter_url or "",
                "company_logo": job.company_logo or "",
                "company_size": job.company_size or "",
                "application_url": job.application_url or "",
                "easy_apply": job.easy_apply,
                "posted_date": job.posted_date.isoformat() if job.posted_date else None,
                "scraped_timestamp": job.scraped_timestamp.isoformat() if job.scraped_timestamp else None,
                "source_type": job.source_type,
                "post_url": job.post_url,
                "post_author_name": job.post_author_name,
                "application_method": job.application_method,
                "application_email": job.application_email,
                "application_platform": job.application_platform,
                "raw_json": job.raw_json,
            })

        try:
            with open(path, mode="w", encoding="utf-8") as f:
                json.dump(json_data, f, indent=4, ensure_ascii=False)
        except PermissionError as exc:
            raise PermissionError(f"Permission denied to write JSON file: {path}") from exc
